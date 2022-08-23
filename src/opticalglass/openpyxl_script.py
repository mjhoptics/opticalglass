#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright © 2020 Michael J. Hayford
"""

.. Created on Thu Dec 24 17:17:36 2020

.. codeauthor: Michael J. Hayford
"""
import warnings

from openpyxl import load_workbook
from openpyxl import worksheet

from opticalglass import ohara
import opticalglass.glass as catgl
from opticalglass import glassfactory as gfact

NBK7 = gfact.create_glass('N-BK7', 'Schott')
tdat_nbk7 = NBK7.transmission_data()


LaK10 = gfact.create_glass('K-LaK10', 'Sumita')

sumita_cat = LaK10.catalog

tdat = LaK10.transmission_data()


fname = 'OHARA.xlsx'
with warnings.catch_warnings():
    warnings.filterwarnings("ignore", category=UserWarning)
    wb = load_workbook(catgl.get_filepath(fname))
ws = wb.active

glass_str = 'Glass '
data_header_offset = 0
glass_name_offset = 1
for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
    try:
        name_col_offset = r.index(glass_str) + 1  # one based
    except ValueError:
        pass
    else:
        glass_header = i
        data_header = i
        # the data headers may be offset from the Glass header row
        data_header += data_header_offset
        break

cols = ws.iter_cols(min_col=name_col_offset,
                    max_col=name_col_offset,
                    values_only=True)
gnames = list(next(cols))
for j, gname in enumerate(gnames[data_header:], start=data_header+1):
    if len(gname) > 0:
        data_start = j
        break

while len(gnames) > 0:
    if gnames[-1] is None:
        gnames.pop()
    elif len(gnames[-1]) == 0:
        gnames.pop()
    else:
        break

num_glasses = len(gnames[data_header:])

for j, c in enumerate(ws.iter_cols(min_col=name_col_offset,
                                   max_col=name_col_offset,
                                   values_only=True), start=1):
    if len(c) > 0:
        data_start = j
        break

cols = ws.iter_cols(min_col=name_col_offset, max_col=name_col_offset,
                    min_row=data_start, max_row=ws.max_row, values_only=True)
gnames = list(cols)
while len(gnames) > 0 and len(gnames[-1]) == 0:
    gnames.pop()

# col = ws.iter_cols(min_col=name_col_offset, max_col=name_col_offset,
#                    min_row=1, max_row=num_glasses, values_only=True)
ws.title
print(wb.sheetnames)
ws.cell(2,3)
ws.cell(3,3)
ws.cell(3,5)
c35=ws.cell(3,5)
c35.value
ws.max_row
#len(tuple(ws.rows))

# offsets for Ohara
#data_header = 1
#data_start = 2+1
num_glasses = 134
#name_col_offset = 1+1
#coef_col_offset = 60
#index_col_offset = 4

ws.dimensions
ws.max_row
ws.max_column

cols = ws.iter_cols(min_col=name_col_offset, max_col=name_col_offset,
                    min_row=data_start, max_row=num_glasses, values_only=True)

name_col = next(cols)
# for i, n in enumerate(name_col):
#     print(i, n)

ohara_cat = ohara.OharaCatalog()
ohara_gnames = ohara_cat.get_glass_names()
for i, n in enumerate(ohara_gnames):
    print(i, n)

