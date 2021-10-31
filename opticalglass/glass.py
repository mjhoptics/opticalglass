#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright © 2019 Michael J. Hayford
""" Support for Glass catalogs and instances

The ``glass`` module contains the two base classes fundamental to the
:mod:`opticalglass` module. The :class:`~opticalglass.glass.GlassCatalog` class
implements as much of the common functionality needed for access to the
catalog data as possible.

The :class:`~opticalglass.glass.Glass` is an interface to the data for a particular
glass in a catalog. The primary function of interest for optical calculations
is :func:`~opticalglass.glass.Glass.rindex` which returns the refractive index at the
input wavelength (nm).

A factory interface to ``Glass`` creation is the function
:func:`~opticalglass.glassfactory.create_glass` that returns a ``Glass``
instance of the appropriate catalog type, given the glass and catalog names.

.. codeauthor: Michael J. Hayford
"""
import itertools
import logging
import warnings
from pathlib import Path

import xlrd
from openpyxl import load_workbook

import numpy as np
import pandas as pd
from scipy import linalg

from . import buchdahl
from . import glasserror as ge
from .util import Singleton, Counter
from .spectral_lines import get_wavelength

from .caselessDictionary import CaselessDictionary


def get_filepath(fname):
    """ given a (spreadsheet) file name, return a complete Path to the file

    The data files included with the ``opticalglass`` package are located in
    a data directory in the package hierarchy.
    ::

        opticalglass/
            opticalglass/
                data/
                    fname

    Args:
        fname (str): the spreadsheet filename, including extender

    Returns:
        str: full path including filename for requested spreadsheet
    """
    pth = Path(__file__).resolve().parent
    return pth/'data'/fname


def calc_glass_constants(nd, nF, nC, *partials):
    """Given central, blue and red refractive indices, calculate Vd and PFd.
    
    Args:
        nd, nF, nC: refractive indices at central, short and long wavelengths
        partials (tuple): if present, 2 ref indxs, n4 and n5, wl4 < wl5
        
    Returns:
        V-number and relative partial dispersion from F to d

    If `partials` is present, the return values include the central wavelength
    index and the relative partial dispersion between the 2 refractive indices
    provided from `partials`.
    """
    dFC = nF-nC
    vd = (nd - 1.0)/dFC
    PFd = (nF-nd)/dFC
    if len(partials) == 2:
        n4, n5 = partials
        P45 = (n4-n5)/dFC
        return nd, vd, PFd, P45
    return vd, PFd


def calc_buchdahl_coords(nd, nF, nC, wlns=('d', 'F', 'C'),
                         ctype=None, **kwargs):
    """Given central, blue and red refractive indices, calculate the Buchdahl
    chromatic coefficients.

    Arguments:
        nd: central refractive index
        nF: "blue" refractive index
        nC: "red" refractive index
        wlns: wavelengths for the 3 refractive indices
        ctype: if "disp_coefs", return dispersion coefficients, otherwise the
               quadratic coefficients
    """
    wv0 = buchdahl.get_wv(wlns[0])

    omF = buchdahl.omega(buchdahl.get_wv(wlns[1]) - wv0)
    omC = buchdahl.omega(buchdahl.get_wv(wlns[2]) - wv0)

    a = np.array([[omF, omF**2], [omC, omC**2]])
    b = np.array([nF-nd, nC-nd])
    coefs = np.linalg.solve(a, b)
    if ctype == "disp_coefs":
        coefs /= (nd - 1)
    return nd, coefs


def fit_buchdahl_coords(indices, degree=2,
                        wlns=['d', 'h', 'g', 'F', 'e', 'C', 'r']):
    """Given central, 4 blue and 2 red refractive indices, do a least squares
    fit for the Buchdahl chromatic coefficients.
    """
    rind0 = indices[0]
    wv0 = buchdahl.get_wv(wlns[0])
    om = [buchdahl.omega(buchdahl.get_wv(w) - wv0) for w in wlns]

    a = np.array([[o**(i+1) for i in range(degree)] for o in om])
    b = np.array(indices) - rind0

    results = linalg.lstsq(a, b)
    coefs = results[0]

    return rind0, coefs


def get_glass_map_arrays(cat, d_str, F_str, C_str, **kwargs):
    """ return index and dispersion data arrays for input spectral range

    Args:
        cat: glass catalog instance, source for returned data
        d_str (str): central wavelength string
        F_str (str): blue end wavelength string
        C_str (str): red end wavelength string
        partials (tuple): kwarg if present, 2 wvls, wl4 and wl5, wl4 < wl5

    Returns:
        index, V-number, partial dispersion, Buchdahl coefficients, and
        glass names
    """
    names = cat.get_glass_names()

    nd = cat.df['refractive indices'][d_str].to_numpy(dtype=float)
    nF = cat.df['refractive indices'][F_str].to_numpy(dtype=float)
    nC = cat.df['refractive indices'][C_str].to_numpy(dtype=float)

    nd, coefs = calc_buchdahl_coords(
        nd, nF, nC, wlns=(d_str, F_str, C_str), **kwargs)

    if 'partials' in kwargs:
        wl_a, wl_b = kwargs['partials']
        na = cat.df['refractive indices'][wl_a].to_numpy(dtype=float)
        nb = cat.df['refractive indices'][wl_b].to_numpy(dtype=float)
        nd, vd, PFd, Pab = calc_glass_constants(nd, nF, nC, na, nb)
    else:
        vd, Pab = calc_glass_constants(nd, nF, nC)

    return nd, vd, Pab, coefs[0], coefs[1], names


class GlassCatalogSpreadsheet:
    """ spreadsheet-based glass catalog """

    def get_glass_names(self):
        """ returns a list of glass names """
        pass

    def get_column_names(self, dindex=None, num=None, add_offset=0):
        """ returns a list of column headers """
        pass

    def get_data_for_glass(self, gindex, dindex=None, num=None):
        """ returns an array of glass data for the glass at *gindex*

        Args:
            gindex: glass index into spreadsheet
            dindex: the starting column of the desired data
            num: number of data items (cells) to retrieve

        Returns: list of data items
        """
        pass

    def get_transmission_wvl(self, header_str):
        """Returns the wavelength value from the transmission data header string."""
        pass

    def transmission_data(self, gindex):
        """ returns an array of transmission data for the glass at *gindex*

        Args:
            gindex: glass index into spreadsheet
            transmission_offset: the starting column of the transmission data
            num_wvls: number of wavelengths for the data

        Returns:
            list of wavelength, transmittance pairs
        """
        header = self.get_column_names(
            dindex=self.transmission_offset,
            num=self.num_wvls,
            add_offset=self.transmission_header_offset)

        data = self.get_data_for_glass(gindex,
                                       dindex=self.transmission_offset,
                                       num=self.num_wvls)

        trns_data = []
        for h, d in zip(header, data):
            w = self.get_transmission_wvl(h)
            if d is not None and d != '' and d != ' ' and d != '-':
                trns_data.append((w, d))
        return trns_data


class GlassCatalogXLS(GlassCatalogSpreadsheet):
    """ Older, Excel XLS-based glass catalog

    Args:
        name: name of the glass catalog
        fname: excel filename, located in ``data`` directory
        glass_str: the header string for the Glass column in fname
        coef_str: the header string for the first refractive index coefficient
                  column in fname
        rindex_str: the header string for the first refractive index value
                    column in fname

    Attributes:
        num_glasses: number of glasses in the catalog
        num_columns: number of colums in the spreadsheet
        data_header: the row containing the data header labels
        data_start: first row in the spreadsheet contain glass data
        name_col_offset: the column offset for glass_str
        coef_col_offset: the column offset for coef_str
        num_coefs: the number of coefficients in the index equation
        index_col_offset: the column offset for rindex_str
        data_header_offset: the row offset of the data headers from the
                            glass_str row
        nline_str: a dict of strings mapping header strings used for index for
                   spectral lines. **Must be provided by the derived class**
    """

    def __init__(self, name, fname, glass_str, coef_str, rindex_str,
                 num_coefs=6, data_header_offset=0, glass_name_offset=1,
                 transmission_offset=0, num_wvls=0,
                 transmission_header_offset=0):
        # data_index is zero-based
        self.based = 0
        self.name = name
        # Open the workbook
        xl_workbook = xlrd.open_workbook(get_filepath(fname))
        self.xl_data = xl_workbook.sheet_by_index(0)

        for i in range(0, self.xl_data.nrows):
            try:
                self.name_col_offset = (self.xl_data.row_values(i, 0)
                                        .index(glass_str))
                glass_header = i
                self.data_header = i
                # the data headers may be offset from the Glass header row
                self.data_header += data_header_offset
                break
            except ValueError:
                pass

        for j in range(glass_header+glass_name_offset, self.xl_data.nrows):
            gname = self.xl_data.cell_value(j, self.name_col_offset)
            if len(gname) > 0:
                self.data_start = j
                break

        gnames = self.get_glass_names()
        self.num_glasses = len(gnames)

        colnames = self.xl_data.row_values(self.data_header, 0)
        self.coef_col_offset = colnames.index(coef_str)
        self.num_coefs = num_coefs
        self.index_col_offset = colnames.index(rindex_str)

        # data slice for 10mm transmission
        self.transmission_offset = transmission_offset
        self.num_wvls = num_wvls
        self.transmission_header_offset = transmission_header_offset

        # build an alphabetical list of decoded glass names
        glass_list = [(decode_glass_name(gn), gn, name)
                      for gn in self.get_glass_names()]
        glass_list = sorted(glass_list, key=lambda glass: glass[0][0])
        # build a lookup dict of the glass defs keyed to decoded glass names
        glass_lookup = {gn_decode: (gn, gc)
                        for gn_decode, gn, gc in glass_list}
        # attach these 'static' lists to class variables
        self.__class__.glass_list = glass_list
        self.__class__.glass_lookup = glass_lookup

    def catalog_name(self):
        return self.name

    def get_glass_names(self):
        """ returns a list of glass names """
        gnames = self.xl_data.col_values(self.name_col_offset, self.data_start)
        # filter out any empty cells at the end
        while gnames and len(gnames[-1]) == 0:
            gnames.pop()
        return gnames

    def get_column_names(self, dindex=None, num=None, add_offset=0):
        """ returns a list of column headers """
        if dindex is not None:  # get a partial row of at least 1 item
            num = num if num is not None else 1
            col_start = dindex
            row_start = self.data_header + add_offset
            colnames = self.xl_data.row_values(row_start,
                                               col_start, col_start+num)
        else:  # get the entire row
            row_start = self.data_header + add_offset
            colnames = self.xl_data.row_values(row_start, 0)

        return colnames

    def glass_index(self, gname):
        """ returns the glass index (row) for glass name `gname`

        Args:
            gname (str): glass name

        Returns:
            int: the index (row) of the requested glass

        Raises:
            GlassNotFoundError: if **gname** doesn't match any header string
        """
        gnames = self.xl_data.col_values(self.name_col_offset, self.data_start)
        if gname in gnames:
            gindex = gnames.index(gname)
        else:
            logging.info('glass %s not found in catalog %s', gname, self.name)
            raise ge.GlassNotFoundError(self.name, gname)

        return gindex

    def data_index(self, dname):
        """ returns the data index (column) for data `dname`

        Args:
            dname (str): header string for data

        Returns:
            int: the index (column) of the requested data

        Raises:
            GlassDataNotFoundError: if **dname** doesn't match any header string
        """
        if dname in self.xl_data.row_values(self.data_header, 0):
            dindex = self.xl_data.row_values(self.data_header, 0).index(dname)
        else:
            logging.info('data type %s not found in catalog %s', dname, self.name)
            raise ge.GlassDataNotFoundError(self.name, dname)

        return dindex

    def get_data_for_glass(self, gindex, dindex=None, num=None):
        """ returns an array of glass data for the glass at *gindex*

        Args:
            gindex: glass index into spreadsheet
            dindex: the starting column of the desired data
            num: number of data items (cells) to retrieve

        Returns: list of data items
        """
        if dindex is not None:  # get a partial row of at least 1 item
            num = num if num is not None else 1
            col_start = dindex
            glass_data = self.xl_data.row_values(self.data_start+gindex,
                                                 col_start, col_start+num)
        else:  # get the entire row
            glass_data = self.xl_data.row_values(self.data_start+gindex, 0)

        return glass_data

    def glass_data(self, gindex):
        """ returns an array of data for the glass at **gindex** """
        return self.xl_data.row_values(self.data_start+gindex, 0)

    def catalog_data(self, dindex):
        """ returns an array of data at column **dindex** for all glasses """
        return self.xl_data.col_values(dindex, self.data_start,
                                       self.data_start+self.num_glasses)

    def glass_coefs(self, gindex):
        """ returns an array of glass coefficients for the glass at **gindex** """
        return (self.xl_data.row_values(self.data_start+gindex,
                                        self.coef_col_offset,
                                        self.coef_col_offset+self.num_coefs))

    def get_transmission_wvl(self, header_str):
        """Returns the wavelength value from the transmission data header string."""
        return float(header_str)

    def glass_map_data(self, wvl='d', **kwargs):
        """ return index and dispersion data for all glasses in the catalog

        Args:
            wvl (str): the central wavelength for the data, either 'd' or 'e'

        Returns:
            index, V-number, partial dispersion, Buchdahl coefficients, and
            glass names
        """
        return get_glass_map_arrays(self, wvl, 'F', 'C', **kwargs)


class GlassCatalogXLSX(GlassCatalogSpreadsheet):
    """ Excel XLSX-based glass catalog

    Args:
        name: name of the glass catalog
        fname: excel filename, located in ``data`` directory
        glass_str: the header string for the Glass column in fname
        coef_str: the header string for the first refractive index coefficient
                  column in fname
        rindex_str: the header string for the first refractive index value
                    column in fname

    Attributes:
        num_glasses: number of glasses in the catalog
        num_columns: number of colums in the spreadsheet
        data_header: the row containing the data header labels
        data_start: first row in the spreadsheet contain glass data
        name_col_offset: the column offset for glass_str
        coef_col_offset: the column offset for coef_str
        num_coefs: the number of coefficients in the index equation
        index_col_offset: the column offset for rindex_str
        data_header_offset: the row offset of the data headers from the
                            glass_str row
        nline_str: a dict of strings mapping header strings used for index for
                   spectral lines. **Must be provided by the derived class**
    """

    def __init__(self, name, fname, glass_str, coef_str, rindex_str,
                 num_coefs=6, data_header_offset=0, glass_name_offset=1,
                 transmission_offset=0, num_wvls=0,
                 transmission_header_offset=0):
        # data_index is one-based
        self.based = 1
        self.name = name
        # Open the workbook
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning)
            wb = load_workbook(get_filepath(fname))
        self.ws = ws = wb.active

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            try:
                name_col_offset = row.index(glass_str)+1
            except ValueError:
                pass
            else:
                data_header = i
                # the data headers may be offset from the Glass header row
                data_header += data_header_offset
                self.data_header = data_header
                self.name_col_offset = name_col_offset
                break

        # find beginning and end of data in names column
        cols = ws.iter_cols(min_col=name_col_offset,
                            max_col=name_col_offset,
                            values_only=True)
        gnames = list(next(cols))
        for j, gname in enumerate(gnames[data_header:], start=data_header+1):
            if gname is not None and len(gname) > 0:
                self.data_start = j
                break

        # filter out any empty cells at the end
        while len(gnames) > 0:
            if gnames[-1] is None:
                gnames.pop()
            elif len(gnames[-1]) == 0:
                gnames.pop()
            else:
                break

        self.num_glasses = len(gnames[self.data_start-1:])

        # find items in column header row
        rows = ws.iter_rows(min_row=data_header,
                            max_row=data_header,
                            values_only=True)

        colnames = list(next(rows))

        # filter out any empty cells at the end
        while len(colnames) > 0:
            if colnames[-1] is None:
                colnames.pop()
            elif len(colnames[-1]) == 0:
                colnames.pop()
            else:
                break

        self.num_columns = len(colnames)

        self.coef_col_offset = colnames.index(coef_str)+1
        self.num_coefs = num_coefs
        self.index_col_offset = colnames.index(rindex_str)+1

        # data slice for 10mm transmission
        self.transmission_offset = transmission_offset
        self.num_wvls = num_wvls
        self.transmission_header_offset = transmission_header_offset

        # build an alphabetical list of decoded glass names
        glass_list = [(decode_glass_name(gn), gn, name)
                      for gn in self.get_glass_names()]
        glass_list = sorted(glass_list, key=lambda glass: glass[0][0])
        # build a lookup dict of the glass defs keyed to decoded glass names
        glass_lookup = {gn_decode: (gn, gc)
                        for gn_decode, gn, gc in glass_list}
        # attach these 'static' lists to class variables
        self.__class__.glass_list = glass_list
        self.__class__.glass_lookup = glass_lookup

    def catalog_name(self):
        return self.name

    def get_glass_names(self):
        """ returns a list of glass names """
        cols = self.ws.iter_cols(min_col=self.name_col_offset,
                                 max_col=self.name_col_offset,
                                 min_row=self.data_start,
                                 max_row=self.data_start+self.num_glasses-1,
                                 values_only=True)
        gnames = list(next(cols))
        return gnames

    def get_column_names(self, dindex=None, num=None, add_offset=0):
        """ returns a list of column headers """
        if dindex is not None:  # get a partial row of at least 1 item
            num = num if num is not None else 1
            col_start = dindex
            row_start = self.data_header + add_offset
            rows = self.ws.iter_rows(min_col=col_start,
                                     max_col=col_start+num-1,
                                     min_row=row_start,
                                     max_row=row_start,
                                     values_only=True)
        else:  # get the entire row
            row_start = self.data_header + add_offset
            rows = self.ws.iter_rows(min_col=1,
                                     max_col=self.num_columns,
                                     min_row=row_start,
                                     max_row=row_start,
                                     values_only=True)

        colnames = list(next(rows))
        return colnames

    def glass_index(self, gname):
        """ returns the glass index (row) for glass name `gname`

        Args:
            gname (str): glass name

        Returns:
            int: the 1-based index (row) of the requested glass

        Raises:
            GlassNotFoundError: if **gname** doesn't match any header string
        """
        gnames = self.get_glass_names()
        if gname in gnames:
            gindex = gnames.index(gname)+1
        else:
            logging.info('glass %s not found in catalog %s', gname, self.name)
            raise ge.GlassNotFoundError(self.name, gname)

        return gindex

    def data_index(self, dname):
        """ returns the data index (column) for data `dname`

        Args:
            dname (str): header string for data

        Returns:
            int: the 1-based index (column) of the requested data

        Raises:
            GlassDataNotFoundError: if *dname* doesn't match any header string
        """
        colnames = self.get_column_names()
        if dname in colnames:
            dindex = colnames.index(dname)+1
        else:
            logging.info('data type %s not found in catalog %s',
                         dname, self.name)
            raise ge.GlassDataNotFoundError(self.name, dname)

        return dindex

    def get_data_for_glass(self, gindex, dindex=None, num=None):
        """ returns an array of glass data for the glass at *gindex*

        Args:
            gindex: glass index into spreadsheet
            dindex: the starting column of the desired data
            num: number of data items (cells) to retrieve

        Returns: list of data items
        """
        if dindex is not None:  # get a partial row of at least 1 item
            num = num if num is not None else 1
            col_start = dindex
            row_start = gindex + self.data_start - 1
            rows = self.ws.iter_rows(min_col=col_start,
                                     max_col=col_start+num-1,
                                     min_row=row_start,
                                     max_row=row_start,
                                     values_only=True)
        else:  # get the entire row
            row_start = gindex + self.data_start - 1
            rows = self.ws.iter_rows(min_col=1,
                                     max_col=self.num_columns,
                                     min_row=row_start,
                                     max_row=row_start,
                                     values_only=True)

        glass_data = next(rows)
        return glass_data

    def glass_data(self, gindex):
        """ returns an array of data for the glass at *gindex* """
        glass_data = self.get_data_for_glass(gindex)
        return glass_data

    def glass_coefs(self, gindex):
        """ returns an array of glass coefficients for the glass at *gindex* """
        glass_data = self.get_data_for_glass(gindex,
                                             dindex=self.coef_col_offset,
                                             num=self.num_coefs)
        return glass_data

    def catalog_data(self, dindex):
        """ returns an array of data at column *dindex* for all glasses """
        cols = self.ws.iter_cols(min_col=dindex,
                                 max_col=dindex,
                                 min_row=self.data_start,
                                 max_row=self.data_start+self.num_glasses-1,
                                 values_only=True)
        return list(next(cols))

    def get_transmission_wvl(self, header_str):
        """Returns the wavelength value from the transmission data header string."""
        return float(header_str)

    def glass_map_data(self, wvl='d', **kwargs):
        """ return index and dispersion data for all glasses in the catalog

        Args:
            wvl (str): the central wavelength for the data, either 'd' or 'e'

        Returns:
            index, V-number, partial dispersion, Buchdahl coefficients, and
            glass names
        """
        return get_glass_map_arrays(self, wvl, 'F', 'C', **kwargs)


class GlassExcel:
    """ base optical glass

    Attributes:
        gname: the glass name
        gindex: the index into the glass list
        catalog: the GlassCatalog this glass is associated with.
                 **Must be provided by the derived class**
        coefs: list of coefficients for calculating refractive index vs wv
    """

    def __init__(self, gname):
        self.gindex = self.catalog.glass_index(gname)
        self.gname = gname
        self.coefs = self.catalog.glass_coefs(self.gindex)

    def __str__(self):
        return self.catalog.name + ' ' + self.name() + ': ' + self.glass_code()

    def __repr__(self):
        return "{!s}('{}')".format(type(self).__name__, self.gname)

    def sync_to_restore(self):
        """ hook routine to restore the gindex given gname """
        self.gindex = self.catalog.glass_index(self.gname)
        if not hasattr(self, 'coefs'):
            self.coefs = self.catalog.glass_coefs(self.gindex)

    def glass_code(self, nd_str, vd_str):
        """ returns the 6 digit glass code, combining index and V-number """
        nd = self.glass_item(nd_str)
        vd = self.glass_item(vd_str)
        return str(1000*round((nd - 1), 3) + round(vd/100, 3))

    def glass_data(self):
        """ returns the raw spreadsheet data for the glass """
        return self.catalog.glass_data(self.gindex)

    def name(self):
        """ returns the glass name, :attr:`gname` """
        return self.gname

    def catalog_name(self):
        """ returns the glass name, :attr:`gname` """
        return self.catalog.catalog_name()

    def glass_item(self, dname):
        """ return the value of the **dname** item

        Args:
            dname (str): header string for data

        Returns:
            the *dname* data

        Raises:
            GlassDataNotFoundError: if *dname* doesn't match any header string
        """
        dindex = self.catalog.data_index(dname)
        if dindex is None:
            return None
        else:
            return self.glass_data()[dindex-self.catalog.based]

    def meas_rindex(self, wvl):
        """ returns the measured refractive index at wvl

        Args:
            wvl: a string with a spectral line identifier

        Returns:
            float: the refractive index at wvl

        Raises:
            KeyError: if *wvl* is not in the spectra dictionary
        """
        return self.glass_item(self.catalog.nline_str[wvl])

    def rindex(self, wvl):
        """ returns the interpolated refractive index at wvl

        Args:
            wvl: either the wavelength in nm or a string with a spectral line
                 identifier. for the refractive index query

        Returns:
            float: the refractive index at wv_nm

        Raises:
            KeyError: if ``wvl`` is not in the spectra dictionary
        """
        return self.calc_rindex(get_wavelength(wvl))

    def calc_rindex(self, wv_nm):
        """ returns the interpolated refractive index at wv_nm

        **Must be provided by the derived class**

        Args:
            wv_nm (float or numpy array): wavelength in nm for the refractive index query

        Returns:
            float or numpy array: the refractive index at wv_nm
        """
        pass

    def transmission_data(self):
        """ returns an array of transmission data for the glass

        Returns: list of wavelength, transmittance pairs for 10mm sample
        """
        return self.catalog.transmission_data(self.gindex)

def xl_cols():
    caps_word = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    cols = list(caps_word)
    
    for c1 in caps_word:
        for c2 in caps_word:
            cols.append(c1 + c2)
    return cols


def xl2df(fname):
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning)
        worksheet_df = pd.read_excel(get_filepath(fname), header=None)
    shape = worksheet_df.shape
    worksheet_df.index = pd.RangeIndex(start=1, stop=shape[0]+1, step=1)
    worksheet_df.columns = xl_cols()[:shape[1]]
    return worksheet_df


def df2gcat(xldf, r0, rk, c0, ck, **kwargs):
    name_col_offset = kwargs['name_col_offset']
    # data_header = kwargs['data_header']
    glass_cat = xldf.loc[r0:rk, c0:ck]
    glass_cat.index = xldf.loc[r0:rk, name_col_offset]
    glass_cat.index.name = 'Glass'
    # glass_cat.columns = xldf.loc[xldf.index[data_header], c0:ck]
    return glass_cat


def fill_in_nans(s):
    categories = {}
    data_header = []
    c_last = 'general'
    r1 = rk = 0, s.index[0]
    for i, item in enumerate(s.items()):
        indx, c = item
        if not pd.isna(c):
            if c != c_last:
                if c_last is not None:
                    categories[c_last] = (r1, rk)
                c_last = c
                r1 = i, indx
            rk = i+1, indx
            data_header.append(c)
        else:
            rk = i+1, indx
            data_header.append(c_last)
    return data_header, categories


def gen_data_headers(cat_df, num_rows):
    data_headers = []
    categories = []
    for i in range(num_rows):
        data_header, df_categories = fill_in_nans(cat_df.iloc[i])
        data_headers.append(data_header)
        categories.append(df_categories)
    return data_headers, categories


class GlassCatalogPandasV1(GlassCatalogSpreadsheet):
    """ Pandas-based glass catalog

    Args:
        name: name of the glass catalog
        fname: excel filename, located in ``data`` directory
        glass_str: the header string for the Glass column in fname
        coef_str: the header string for the first refractive index coefficient
                  column in fname
        rindex_str: the header string for the first refractive index value
                    column in fname

    Attributes:
        num_glasses: number of glasses in the catalog
        num_columns: number of colums in the spreadsheet
        data_header: the row containing the data header labels
        data_start: first row in the spreadsheet contain glass data
        name_col_offset: the column offset for glass_str
        coef_col_offset: the column offset for coef_str
        num_coefs: the number of coefficients in the index equation
        index_col_offset: the column offset for rindex_str
        nline_str: a dict of strings mapping header strings used for index for
                   spectral lines. **Must be provided by the derived class**
    """

    def __init__(self, name, fname, *args, **kwargs):
        # data_index is one-based
        self.based = 0
        self.name = name
        # Open the workbook
        worksheet_df = xl2df(fname)
        self.glass_cat = glass_cat = df2gcat(worksheet_df, *args, **kwargs)

        self.gnames = gnames = self.get_glass_names()
        self.num_glasses = len(gnames)

        self.colnames = colnames = self.get_column_names()
        self.num_columns = len(colnames)

        coef_str = kwargs['coef_str']
        self.coef_col_offset = self.glass_cat.columns.get_loc(coef_str)
        self.num_coefs = kwargs['num_coefs']
        rindex_str = kwargs['rindex_str']
        self.index_col_offset = self.glass_cat.columns.get_loc(rindex_str)

        # data slice for 10mm transmission
        self.transmission_offset = kwargs['transmission_offset']
        self.num_wvls = kwargs['num_wvls']
        twv = kwargs['transmission_wvl_shape']
        self.transmission_wvls = worksheet_df.iloc[twv[0]][twv[2]:twv[3]]

        # build an alphabetical list of decoded glass names
        glass_list = [(decode_glass_name(gn), gn, name)
                      for gn in gnames]
        glass_list = sorted(glass_list, key=lambda glass: glass[0][0])
        # build a lookup dict of the glass defs keyed to decoded glass names
        glass_lookup = {gn_decode: (gn, gc)
                        for gn_decode, gn, gc in glass_list}
        # attach these 'static' lists to class variables
        self.__class__.glass_list = glass_list
        self.__class__.glass_lookup = glass_lookup

    def catalog_name(self):
        return self.name

    def get_glass_names(self):
        """ returns a list of glass names """
        return self.glass_cat.index.array

    def get_column_names(self, dindex=None, num=None, add_offset=0):
        """ returns a list of column headers """
        return self.glass_cat.columns.array

    def glass_index(self, gname):
        """ returns the glass index (row) for glass name `gname`

        Args:
            gname (str): glass name

        Returns:
            int: the 0-based index (row) of the requested glass
        """
        return self.glass_cat.index.get_loc(gname)

    def data_index(self, dname):
        """ returns the data index (column) for data `dname`

        Args:
            dname (str): header string for data

        Returns:
            int: the 1-based index (column) of the requested data

        Raises:
            GlassDataNotFoundError: if *dname* doesn't match any header string
        """
        dname_iloc = self.glass_cat.columns.get_loc(dname)
        if isinstance(dname_iloc, int):
            return dname_iloc
        for i, b in enumerate(dname_iloc, start=1):
            if b:
                break
        return i

    def get_data_for_glass(self, gname, dindex=None, num=None):
        """ returns an array of glass data for the glass at *gname*

        Args:
            gname: glass index into spreadsheet
            dindex: the starting column of the desired data
            num: number of data items (cells) to retrieve

        Returns: list of data items
        """
        if isinstance(gname, int):
            gname = self.glass_cat.index[gname]
        glass_data = self.glass_cat.loc[gname]
        if dindex is not None:  # get a partial row of at least 1 item
            num = num if num is not None else 1
            if isinstance(dindex, str):
                dindex = glass_data.index.get_loc(dindex)
            glass_data = glass_data.iloc[dindex:dindex+num]

        return glass_data

    def glass_data(self, gindex):
        """ returns an array of data for the glass at *gindex* """
        glass_data = self.get_data_for_glass(gindex)
        return glass_data

    def glass_coefs(self, gname):
        """ returns an array of glass coefficients for the glass at *gname* """
        glass_data = self.get_data_for_glass(gname,
                                             dindex=self.coef_col_offset,
                                             num=self.num_coefs)
        return glass_data

    def catalog_data(self, dindex):
        """ returns an array of data at column *dindex* for all glasses """
        dname = self.ws.columns[dindex]
        return self.ws[dname]

    def get_transmission_wvl(self, header_str):
        """Returns the wavelength value from the transmission data header string."""
        return float(header_str)

    def transmission_data(self, gname):
        """ returns an array of transmission data for the glass at *gindex*

        Args:
            gindex: glass index into spreadsheet
            transmission_offset: the starting column of the transmission data
            num_wvls: number of wavelengths for the data

        Returns:
            list of wavelength, transmittance pairs
        """
        data = self.get_data_for_glass(gname,
                                       dindex=self.transmission_offset,
                                       num=self.num_wvls)
        trns_data = []
        for h, d in zip(self.transmission_wvls, data):
            w = self.get_transmission_wvl(h)
            if d is not None and not pd.isna(d):
                trns_data.append((w, d))
        return trns_data

    def glass_map_data(self, wvl='d', **kwargs):
        """ return index and dispersion data for all glasses in the catalog

        Args:
            wvl (str): the central wavelength for the data, either 'd' or 'e'

        Returns:
            index, V-number, partial dispersion, Buchdahl coefficients, and
            glass names
        """
        return get_glass_map_arrays(self, wvl, 'F', 'C', **kwargs)


class GlassV1:
    """ base optical glass, for use with pandas

    Attributes:
        gname: the glass name
        gindex: the index into the glass list
        catalog: the GlassCatalog this glass is associated with.
                 **Must be provided by the derived class**
        coefs: list of coefficients for calculating refractive index vs wv
    """

    def __init__(self, gname):
        # self.gindex = self.catalog.glass_index(gname)
        self.gname = gname
        self.coefs = self.catalog.glass_coefs(gname)

    def __str__(self):
        return self.catalog.name + ' ' + self.name() + ': ' + self.glass_code()

    def __repr__(self):
        return "{!s}('{}')".format(type(self).__name__, self.gname)

    def sync_to_restore(self):
        """ hook routine to restore coefs, if needed """
        # self.gindex = self.catalog.glass_index(self.gname)
        if not hasattr(self, 'coefs'):
            self.coefs = self.catalog.glass_coefs(self.gindex)

    def glass_code(self, nd_str, vd_str):
        """ returns the 6 digit glass code, combining index and V-number """
        nd = self.glass_item(nd_str)
        vd = self.glass_item(vd_str)
        return str(1000*round((nd - 1), 3) + round(vd/100, 3))

    def glass_data(self):
        """ returns the raw spreadsheet data for the glass as a Series """
        return self.catalog.glass_cat.loc[self.gname]

    def name(self):
        """ returns the glass name, :attr:`gname` """
        return self.gname

    def catalog_name(self):
        """ returns the glass name, :attr:`gname` """
        return self.catalog.catalog_name()

    def glass_item(self, dname):
        """ return the value of the **dname** item

        Args:
            dname (str): header string for data

        Returns:
            the *dname* data

        Raises:
            GlassDataNotFoundError: if *dname* doesn't match any header string
        """
        return self.glass_data()[dname]

    def meas_rindex(self, wvl):
        """ returns the measured refractive index at wvl

        Args:
            wvl: a string with a spectral line identifier

        Returns:
            float: the refractive index at wvl

        Raises:
            KeyError: if *wvl* is not in the spectra dictionary
        """
        return self.glass_item(self.catalog.nline_str[wvl])

    def rindex(self, wvl):
        """ returns the interpolated refractive index at wvl

        Args:
            wvl: either the wavelength in nm or a string with a spectral line
                 identifier. for the refractive index query

        Returns:
            float: the refractive index at wv_nm

        Raises:
            KeyError: if ``wvl`` is not in the spectra dictionary
        """
        return self.calc_rindex(get_wavelength(wvl))

    def calc_rindex(self, wv_nm):
        """ returns the interpolated refractive index at wv_nm

        **Must be provided by the derived class**

        Args:
            wv_nm (float or numpy array): wavelength in nm for the refractive index query

        Returns:
            float or numpy array: the refractive index at wv_nm
        """
        pass

    def transmission_data(self):
        """ returns an array of transmission data for the glass

        Returns: list of wavelength, transmittance pairs for 10mm sample
        """
        return self.catalog.transmission_data(self.gname) 


standard_categories = dict(
    refractive_indices = 'refractive indices',
    dispersion_coefficients = 'dispersion coefficients',
    internal_transmission_10 = 'internal transmission mm, 10',
   )
  
    
def replace_category_headers(data_headers, worksheet_df, row, old_keys, new_keys):
    for key in new_keys.keys():
        # span = categories_by_row[row][old_keys[key]]
        # s1, sk = span[0][0], span[1][0]
        cols = worksheet_df.columns
        row, _, col1, colk = old_keys[key]
        s1, sk = cols.get_loc(col1), cols.get_loc(colk)+1
        for i in range(s1, sk):
            data_headers[row][i] = new_keys[key]


def build_glass_cat_index(cat, worksheet_df, *args, **kwargs):
    num_rows, category_row , header_row, data_col = args
    refractive_indices = kwargs['refractive_indices']
    dispersion_coefficients = kwargs['dispersion_coefficients']
    internal_transmission_10 = kwargs['internal_transmission_10']

    # setup the data headers for the glass catalog dataframe
    data_headers, categories_by_row = gen_data_headers(worksheet_df, num_rows)
    mindx0 = pd.MultiIndex.from_arrays([data_headers[category_row][:], 
                                        data_headers[header_row][:]], 
                                       names=['category', 'data item'])

    # some spreadsheets have duplicate column headers; get rid of those
    if mindx0.has_duplicates:
        dups = mindx0.duplicated()
        dups1 = [worksheet_df.columns[i] for i, b in enumerate(dups) if b == True]
        worksheet_df.drop(columns=dups1, inplace=True)

    # generate data headers again; should be all unique
    data_headers, categories_by_row = gen_data_headers(worksheet_df, num_rows)
    cols = worksheet_df.columns

    # data slice for measured/calculated refractive index data
    # spanrn = categories_by_row[category_row][refractive_indices]
    # s1, sk = spanrn[0][0], spanrn[1][0]
    row, _, col1, colk = kwargs['refractive_indices']
    s1, sk = cols.get_loc(col1), cols.get_loc(colk)+1
    refractive_index_wvls = [h.split('n')[-1]
                             for h in data_headers[row][s1:sk]]
    data_headers[header_row][s1:sk] = refractive_index_wvls

    # data slice for 10mm transmission
    # span10 = categories_by_row[category_row][internal_transmission_10]
    # s1, sk = span10[0][0], span10[1][0]
    row, _, col1, colk = kwargs['internal_transmission_10']
    s1, sk = cols.get_loc(col1), cols.get_loc(colk)+1

    def get_transmission_wvl_fct(header_str):
        """Returns the wavelength value from the transmission data header string."""
        return float(header_str)

    get_transmission_wvl = kwargs['get_transmission_wvl']
    get_transmission_wvl = (get_transmission_wvl_fct 
                            if get_transmission_wvl is not None 
                            else kwargs['get_transmission_wvl'])
    transmission_wvl = [get_transmission_wvl(h)
                        for h in data_headers[row][s1:sk]]
    data_headers[header_row][s1:sk] = transmission_wvl

    # replace vendor specific categories with standard ones
    replace_category_headers(data_headers, worksheet_df, category_row,
                             kwargs, standard_categories)

    mindx = pd.MultiIndex.from_arrays([data_headers[category_row][data_col:],
                                       data_headers[header_row][data_col:]],
                                      names=['category', 'data item'])

    # make a new df from the data area of the worksheet
    r0, rk, c0, ck = kwargs['data_extent']
    glass_cat = worksheet_df.loc[r0:rk, c0:ck]
    glass_cat.index = worksheet_df.loc[r0:rk, kwargs['name_col_offset']]
    glass_cat.index.name = 'glass'
    glass_cat.columns = mindx

    return glass_cat


def gen_data_headers_only(cat_df, num_rows):
    data_headers = []
    for i in range(num_rows):
        data_header = [item[1] for item in cat_df.iloc[i].items()]
        data_headers.append(data_header)
    return data_headers


def build_glass_cat_orig(worksheet_df, series_mappings, item_mappings, 
                         *args, **kwargs):
    num_rows, category_row , header_row, data_col = args

    # # setup the data headers for the glass catalog dataframe
    # data_headers, categories_by_row = gen_data_headers(worksheet_df, num_rows)
    # mindx0 = pd.MultiIndex.from_arrays([data_headers[category_row][:], 
    #                                     data_headers[header_row][:]], 
    #                                    names=['category', 'data item'])

    # # some spreadsheets have duplicate column headers; get rid of those
    # if mindx0.has_duplicates:
    #     dups = mindx0.duplicated()
    #     dups1 = [worksheet_df.columns[i] for i, b in enumerate(dups) if b == True]
    #     worksheet_df.drop(columns=dups1, inplace=True)

    # generate data headers again; should be all unique
    data_headers = gen_data_headers_only(worksheet_df, num_rows)
    cols = worksheet_df.columns

    for m in series_mappings:
        category, fct, row, col1, colk = m
        s1, sk = cols.get_loc(col1), cols.get_loc(colk)+1
        for si, h_item in enumerate(data_headers[row][s1:sk], start=s1):
            item = h_item if fct is None else fct(h_item)
            data_headers[header_row][si] = item
            data_headers[category_row][si] = category

    for m in item_mappings:
        category, hdr, row, col = m
        s = cols.get_loc(col)
        item = data_headers[row][s]
        if callable(hdr):
            h_item = hdr(item)
        elif hdr is not None:
            h_item = hdr
        else:
            h_item = item
        data_headers[header_row][si] = h_item
        data_headers[category_row][si] = category

    mindx = pd.MultiIndex.from_arrays([data_headers[category_row][data_col:],
                                       data_headers[header_row][data_col:]],
                                      names=['category', 'data item'])

    # make a new df from the data area of the worksheet
    r0, rk, c0, ck = kwargs['data_extent']
    glass_cat = worksheet_df.loc[r0:rk, c0:ck]
    glass_cat.index = worksheet_df.loc[r0:rk, kwargs['name_col_offset']]
    glass_cat.index.name = 'glass'
    glass_cat.columns = mindx

    return glass_cat


def build_glass_cat(worksheet_df, series_mappings, item_mappings, 
                    *args, **kwargs):
    num_rows, category_row , header_row, data_col = args

    # generate data headers again; should be all unique
    headers = worksheet_df.loc[1:num_rows].copy().astype('object')

    for m in series_mappings:
        category, action, row, col1, colk = m
        for idx in headers.loc[row].loc[col1:colk].index:
            if action is not None:
                headers.at[header_row, idx] = action(headers.loc[row, idx])
            else:
                headers.at[header_row, idx] = headers.loc[row, idx]
            headers.at[category_row, idx] = category

    for m in item_mappings:
        category, hdr, row, col = m
        item = headers.loc[row, col]
        if callable(hdr):
            h_item = hdr(item)
        elif hdr is not None:
            h_item = hdr
        else:
            h_item = item
        headers.at[header_row, col] = h_item
        headers.at[category_row, col] = category

    mindx = pd.MultiIndex.from_arrays([headers.loc[category_row, data_col:],
                                       headers.loc[header_row, data_col:]],
                                      names=['category', 'data item'])

    # make a new df from the data area of the worksheet
    r0, rk, c0, ck = kwargs['data_extent']
    glass_cat = worksheet_df.loc[r0:rk, c0:ck]
    glass_cat.index = worksheet_df.loc[r0:rk, kwargs['name_col_offset']]
    glass_cat.index.name = 'glass'
    glass_cat.columns = mindx

    # weed out any duplicate or empty columns
    if glass_cat.columns.has_duplicates:
        dups = glass_cat.columns.duplicated()
        dups1 = [glass_cat.columns[i] for i, b in enumerate(dups) if b == True]
        glass_cat.drop(columns=dups1, inplace=True)

    # infer dtypes so that downstream numpy routines don't complain about object data
    glass_cat = glass_cat.convert_dtypes()

    return glass_cat

class GlassCatalogPandas(GlassCatalogSpreadsheet):
    """ Pandas-based glass catalog

    Args:
        name: name of the glass catalog
        fname: excel filename, located in ``data`` directory
        glass_str: the header string for the Glass column in fname
        coef_str: the header string for the first refractive index coefficient
                  column in fname
        rindex_str: the header string for the first refractive index value
                    column in fname

    Attributes:
        num_glasses: number of glasses in the catalog
        num_columns: number of colums in the spreadsheet
        data_header: the row containing the data header labels
        data_start: first row in the spreadsheet contain glass data
        name_col_offset: the column offset for glass_str
        coef_col_offset: the column offset for coef_str
        num_coefs: the number of coefficients in the index equation
        index_col_offset: the column offset for rindex_str
        nline_str: a dict of strings mapping header strings used for index for
                   spectral lines. **Must be provided by the derived class**
    """

    # def __init__(self, name, fname, *args, **kwargs):
    def __init__(self, name, fname, series_mappings, item_mappings, 
                 *args, **kwargs):

        self.name = name
        # Open the workbook
        worksheet_df = xl2df(fname)
        self.df = build_glass_cat(worksheet_df, series_mappings, item_mappings, 
                                  *args, **kwargs)
        # self.df = build_glass_cat_index(self, worksheet_df, *args, **kwargs)

        # build an alphabetical list of decoded glass names
        gnames = self.df.index.array
        glass_list = [(decode_glass_name(gn), gn, name)
                      for gn in gnames]
        glass_list = sorted(glass_list, key=lambda glass: glass[0][0])
        # build a lookup dict of the glass defs keyed to decoded glass names
        glass_lookup = {gn_decode: (gn, gc)
                        for gn_decode, gn, gc in glass_list}
        # attach these 'static' lists to class variables
        self.__class__.glass_list = glass_list
        self.__class__.glass_lookup = glass_lookup

    def catalog_name(self):
        return self.name

    def get_glass_names(self):
        """ returns a list of glass names """
        return self.df.index.array

    def get_column_names(self, dindex=None, num=None, add_offset=0):
        """ returns a list of column headers """
        return self.df.columns.array

    def glass_index(self, gname):
        """ returns the glass index (row) for glass name `gname`

        Args:
            gname (str): glass name

        Returns:
            int: the 0-based index (row) of the requested glass
        """
        return self.df.index.get_loc(gname)

    def data_index(self, dname):
        """ returns the data index (column) for data `dname`

        Args:
            dname (str): header string for data

        Returns:
            int: the 1-based index (column) of the requested data

        Raises:
            GlassDataNotFoundError: if *dname* doesn't match any header string
        """
        dname_iloc = self.df.columns.get_loc(dname)
        if isinstance(dname_iloc, int):
            return dname_iloc
        for i, b in enumerate(dname_iloc, start=1):
            if b:
                break
        return i

    def get_data_for_glass(self, gname, dindex=None, num=None):
        """ returns an array of glass data for the glass at *gname*

        Args:
            gname: glass index into spreadsheet
            dindex: the starting column of the desired data
            num: number of data items (cells) to retrieve

        Returns: list of data items
        """
        if isinstance(gname, int):
            gname = self.df.index[gname]
        glass_data = self.df.loc[gname]
        if dindex is not None:  # get a partial row of at least 1 item
            num = num if num is not None else 1
            if isinstance(dindex, str):
                dindex = glass_data.index.get_loc(dindex)
            glass_data = glass_data.iloc[dindex:dindex+num]

        return glass_data

    def glass_data(self, gindex):
        """ returns an array of data for the glass at *gindex* """
        glass_data = self.get_data_for_glass(gindex)
        return glass_data

    def glass_coefs(self, gname):
        """ returns an array of glass coefficients for the glass at *gname* """
        glas = self.df.loc[gname]
        coefs = glas['dispersion coefficients'].to_numpy(dtype=float)
        return coefs

    def catalog_data(self, dindex):
        """ returns an array of data at column *dindex* for all glasses """
        dname = self.ws.columns[dindex]
        return self.ws[dname]

    def transmission_data(self, gname):
        """ returns an array of transmission data for the glass at *gindex*

        Args:
            gindex: glass index into spreadsheet
            transmission_offset: the starting column of the transmission data
            num_wvls: number of wavelengths for the data

        Returns:
            list of wavelength, transmittance pairs
        """
        gla = self.df.loc[gname]
        return gla['internal transmission mm, 10'].array

    def glass_map_data(self, wvl='d', **kwargs):
        """ return index and dispersion data for all glasses in the catalog

        Args:
            wvl (str): the central wavelength for the data, either 'd' or 'e'

        Returns:
            index, V-number, partial dispersion, Buchdahl coefficients, and
            glass names
        """
        return get_glass_map_arrays(self, wvl, 'F', 'C', **kwargs)


class Glass:
    """ base optical glass, for use with pandas

    Attributes:
        gname: the glass name
        catalog: the GlassCatalog this glass is associated with.
                 **Must be provided by the derived class**
        coefs: list of coefficients for calculating refractive index vs wv
    """

    def __init__(self, gname):
        self.gname = gname
        self.coefs = self.catalog.glass_coefs(gname)

    def __str__(self):
        return self.catalog.name + ' ' + self.name() + ': ' + self.glass_code()

    def __repr__(self):
        return "{!s}('{}')".format(type(self).__name__, self.gname)

    def sync_to_restore(self):
        """ hook routine to restore coefs, if needed """
        if not hasattr(self, 'coefs'):
            self.coefs = self.catalog.glass_coefs(self.gname)

    def glass_code(self, d_str='d', vd_str='vd'):
        """ returns the 6 digit glass code, combining index and V-number """
        glas = self.glass_data()
        nd = glas['refractive indices'][d_str]
        vd = glas['abbe number'][vd_str]
        return str(1000*round((nd - 1), 3) + round(vd/100, 3))

    def glass_data(self):
        """ returns the raw spreadsheet data for the glass as a Series """
        return self.catalog.df.loc[self.gname]

    def name(self):
        """ returns the glass name, :attr:`gname` """
        return self.gname

    def catalog_name(self):
        """ returns the glass name, :attr:`gname` """
        return self.catalog.catalog_name()

    def meas_rindex(self, wvl):
        """ returns the measured refractive index at wvl

        Args:
            wvl: a string with a spectral line identifier

        Returns:
            float: the refractive index at wvl

        Raises:
            KeyError: if *wvl* is not in the spectra dictionary
        """
        rindx = self.glass_data()['refractive indices'][wvl]
        return rindx

    def rindex(self, wvl):
        """ returns the interpolated refractive index at wvl

        Args:
            wvl: either the wavelength in nm or a string with a spectral line
                 identifier. for the refractive index query

        Returns:
            float: the refractive index at wv_nm

        Raises:
            KeyError: if ``wvl`` is not in the spectra dictionary
        """
        return self.calc_rindex(get_wavelength(wvl))

    def calc_rindex(self, wv_nm):
        """ returns the interpolated refractive index at wv_nm

        **Must be provided by the derived class**

        Args:
            wv_nm (float or numpy array): wavelength in nm for the refractive index query

        Returns:
            float or numpy array: the refractive index at wv_nm
        """
        pass

    def transmission_data(self):
        """ returns an array of transmission data for the glass

        Returns: list of wavelength, transmittance pairs for 10mm sample
        """
        glas = self.glass_data()
        return glas['internal transmission mm, 10'].to_numpy(dtype=float)


def decode_glass_name(glass_name):
    """Split glass_name into prefix, group, num, suffix.

    Manufacturers glass names follow a common pattern. At the simplest, it is
    a short character string, typically used to identify a particular glass
    composition, with a numeric qualifier. The composition group and product
    serial number are combined to form the basic product id, the group_num:

        - F2
        - SF56

    Manufacturers will often use a single character prefix to indicate
    different categories of glasses, e.g. moldable or "New":

        - N-BK7
        - P-LASF50

    Similarly, a suffix with one or more characters is often used to
    differentiate between different variations of the same base material.

        - N-SF57
        - N-SF57HT
        - N-SF57HTultra

    This function takes an input glass name and returns a tuple of strings. A
    valid glass_name should always have a non-null group_num; prefixes and
    suffixes are optional and used differently by different manufacturers.

        * group_num, prefix, suffix
        * group, num = group_num

    Args:
        glass_name (str): a glass manufacturer's glass name

    Returns: group_num, prefix, suffix, where group_num = group, num

    Returned strings are uppercase.

    """
    gn = glass_name.upper().split('-')
    suffix = ''
    if len(gn) == 1:
        prefix = ''
        gn2 = gn[0]
    elif len(gn) == 3:
        prefix = gn[0]
        suffix = gn[2]
        gn2 = gn[1]
    elif len(gn) == 2:
        if len(gn[0]) < 3:
            prefix = gn[0]
            gn2 = gn[1]
        else:
            prefix = ''
            gn2 = gn[0]
            suffix = gn[1]

    group = gn2
    num = ''
    for i, char in enumerate(gn2):
        if char.isdigit():
            start = i
            while i < len(gn2) and gn2[i].isdigit():
                i += 1
            group = gn2[:start].rstrip()
            num = gn2[start:i]
            break
    suffix = gn2[i:] if suffix == '' else suffix
    group_num = group, num
    return group_num, prefix, suffix


def glass_catalog_stats(glass_list, do_print=False):
    """Decode all of the glass names in glass_cat_name.

    Print out the original glass names and the decoded version side by side.

    Args:
        glass_list: ((group, num), prefix, suffix), glass_name, glass_cat_name
        do_print (bool): if True, print the glass name and the decoded version

    Returns:
        groups (dict): all catalog groups
        group_num (dict): all glasses with multiple variants
        prefixes (dict): all the non-null prefixes used
        suffixes (dict): all the non-null suffixes used
    """
    groups = Counter()
    group_nums = Counter()
    prefixes = Counter()
    suffixes = Counter()
    for g in glass_list:
        (group_num, prefix, suffix), gn, gc = g
        group, num = group_num
        if prefix != '':
            prefixes[prefix] += 1
        if suffix != '':
            suffixes[suffix] += 1
        groups[group] += 1
        group_nums[group_num] += 1
        if do_print:
            fmt_3 = "{:14s} {:>2s}  {:8s}  {:12s}"
            print(fmt_3.format(gn, prefix, group+num, suffix))
    # filter group_nums to include only multi-use items
    group_nums = {k: v for k, v in group_nums.items() if v > 1}
    return groups, group_nums, prefixes, suffixes


class Robb1983Catalog(metaclass=Singleton):
    """ glass catalog based on data in Robb, et als 1983 paper on Buchdahl's
    chromatic coordinate

    The data used by this class is from the 1983 paper by Paul N. Robb and
    R. I. Mercado, `Calculation of refractive indices using Buchdahl’s
    chromatic coordinate <https://doi.org/10.1364/AO.22.001198>`_ . The
    copyright date for all 5 catalogs cited was 1980.

    Args:
        fname: filename, located in ``data`` directory

    Attributes:
        glass_db: dict lookup by catalog and glass name, value is decoded
                  glassname and Buchdahl coefficients
        glass_list: list of decoded_glassname, glassname, and catalog per glass

    """

    _cat_names = ["SCHOTT", "OHARA", "HOYA", "CORNING-FRANCE", "CHANCE"]

    def __init__(self, fname='robb1983_data_final.txt'):
        # Open the workbook
        glass_db = CaselessDictionary()
        glass_list = []
        rndx_list = []
        nu1_list = []
        nu2_list = []
        with get_filepath(fname).open() as f_input:
            for line in f_input:
                if line[0] == '#':
                    if 'GLASS CATALOGUE' in line:
                        tokens = line[1:].split()
                        catalog = 'Robb1983.' + tokens[0]
                        glass_db[catalog] = {}
                else:
                    tokens = line.split()
                    gname = tokens[0]
                    try:
                        gname_decode = decode_glass_name(gname)
                    except UnboundLocalError:
                        print(catalog, gname)
                    else:
                        rndx = float(tokens[1])
                        nu1 = float(tokens[2])
                        nu2 = float(tokens[3])
                        glass_db[catalog][gname] = (
                            gname_decode, rndx, nu1, nu2)
                        glass_list.append((gname_decode, gname, catalog))
                        rndx_list.append(rndx)
                        nu1_list.append(nu1)
                        nu2_list.append(nu2)

        self.glass_db = glass_db
        self.glass_list = glass_list
        self._glass_data = CaselessDictionary()

    @property
    def glass_data(self):
        if len(self._glass_data) == 0:
            for rbk, rbv in self.glass_db.items():
                gnames = list(rbv.keys())
                gdata = np.array([[d[1], d[2], d[3]] for d in rbv.values()])
                self._glass_data[rbk] = gnames, gdata
        return self._glass_data

    def create_glass(self, gname, gcat):
        catalog = gcat if 'Robb1983.' in gcat else 'Robb1983.' + gcat
        try:
            gdata = self.glass_db[catalog][gname]
        except KeyError:
            return None
        else:
            wv0 = buchdahl.get_wv('d')
            gname_decode, rndx, nu1, nu2 = gdata
            g = buchdahl.Buchdahl(wv0, rndx, (nu1, nu2), mat=gname, cat=gcat)
            return g

    def catalog_name(self):
        return 'Robb1983'

    def get_glass_names(self, gcat=None):
        """ returns a list of glass names """
        if gcat is not None:
            return self.glass_data[gcat][0]
        else:  # return all of the catalogs' names
            gnames = [gn[0] for gn in self.glass_data.values()]
            return list(itertools.chain.from_iterable(gnames))

    def glass_map_data(self, wvl='d', **kwargs):
        """ return index and dispersion data for all glasses in the catalog

        Args:
            wvl (str): the central wavelength for the data, either 'd' or 'e'

        Returns:
            index, V-number, partial dispersion, Buchdahl coefficients, and
            glass names
        """
        return self.get_glass_map_arrays(wvl, 'F', 'C', **kwargs)

    def get_glass_map_arrays(self, d_str, F_str, C_str, **kwargs):
        """ return index and dispersion data arrays for input spectral range

        Args:
            nd_str (str): central wavelength string
            nf_str (str): blue end wavelength string
            nc_str (str): red end wavelength string

        Returns:
            index, V-number, partial dispersion, Buchdahl coefficients, and
            glass names
        """
        wv_0 = buchdahl.get_wv('d')
        wv_d = buchdahl.get_wv(d_str)
        wv_F = buchdahl.get_wv(F_str)
        wv_C = buchdahl.get_wv(C_str)

        om_d = buchdahl.omega(wv_d - wv_0)
        om_F = buchdahl.omega(wv_F - wv_0)
        om_C = buchdahl.omega(wv_C - wv_0)

        gnames, gdata = self.glass_data[kwargs['cat_name']]

        omm_d = np.array([om_d, om_d**2])
        nd = np.matmul(gdata[:, 1:], omm_d) + gdata[:, 0]

        omm_F = np.array([om_F, om_F**2])
        nF = np.matmul(gdata[:, 1:], omm_F) + gdata[:, 0]

        omm_C = np.array([om_C, om_C**2])
        nC = np.matmul(gdata[:, 1:], omm_C) + gdata[:, 0]

        vd, PCd = calc_glass_constants(nd, nF, nC)

        coefs = np.array(gdata[:, 1:3])
        ctype = kwargs.get('ctype', None)
        if ctype == "disp_coefs":
            coefs[:, 0] /= (nd - 1)
            coefs[:, 1] /= (nd - 1)

        return nd, vd, PCd, coefs[:, 0], coefs[:, 1], gnames


_robb1983_list = CaselessDictionary({
        'Robb1983.Schott': Robb1983Catalog(),
        'Robb1983.Ohara': Robb1983Catalog(),
        'Robb1983.Hoya': Robb1983Catalog(),
        'Robb1983.Corning-France': Robb1983Catalog(),
        'Robb1983.Chance': Robb1983Catalog(),
        })
